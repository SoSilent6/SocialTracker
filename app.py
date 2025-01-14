import os
from flask import Flask, render_template, jsonify
import pandas as pd
from datetime import datetime
import math
import openpyxl

app = Flask(__name__)

def format_number(value):
    """Format numbers with comma separators"""
    try:
        if pd.isna(value) or value == 0:
            return ""
        if isinstance(value, (int, float)):
            return f"{int(value):,}"
        return value
    except:
        return value

def format_percentage(value):
    """Format percentage values"""
    try:
        if pd.isna(value) or value == 0:
            return ""
        if isinstance(value, str) and ('#VALUE!' in value or '#DIV/0!' in value):
            return ""
        return f"{float(value):.2f}%"
    except:
        return ""

@app.route('/')
def show_excel():
    try:
        # Load Excel file
        wb = openpyxl.load_workbook('Social Follower Count.xlsx', data_only=True)
        sheet = wb['Info']
        
        # Convert to DataFrame
        data = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2):
            row_data = [cell.value if cell.value is not None else "" for cell in row]
            if any(row_data):
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        
        # Create result DataFrame with default columns
        result_df = pd.DataFrame({
            "代币": df['Token'].apply(lambda x: f'<div style="display: flex; align-items: center;"><a href="/token/{x}" class="详细数据-btn">详细数据</a><span style="margin-left: 10px">{x}</span></div>'),
            "X粉丝数量": df['X粉丝数量'].apply(format_number),
            "24h涨跌(X)": df['24h粉丝涨跌(X)'].apply(format_percentage),
            "Discord粉丝数量": df['Discord粉丝数量'].apply(format_number),
            "24h涨跌(Discord)": df['24h粉丝涨跌(Discord)'].apply(format_percentage),
            "Telegram粉丝数量": df['Telegram粉丝数量'].apply(format_number),
            "24h涨跌(Telegram)": df['24h粉丝涨跌(Telegram)'].apply(format_percentage),
            # Hidden columns (available through DIY)
            "3天涨跌(X)": df['3天粉丝涨跌(X)'].apply(format_percentage),
            "3天涨跌(Discord)": df['3天粉丝涨跌(Discord)'].apply(format_percentage),
            "3天涨跌(Telegram)": df['3天粉丝涨跌(Telegram)'].apply(format_percentage)
        })
        
        # Calculate last update time
        mod_timestamp = os.path.getmtime('Social Follower Count.xlsx')
        time_diff = (datetime.now().timestamp() - mod_timestamp) / 60
        last_update = f"{math.floor(time_diff)}分钟前" if time_diff < 60 else f"{math.floor(time_diff / 60)}小时前"
        
        # Convert DataFrame to dict for JSON serialization
        data_dict = result_df.to_dict('records')
        
        return render_template('index.html', 
                             data=data_dict,
                             last_update=last_update)
    
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/token/<token_name>')
def token_detail(token_name):
    try:
        # Load Excel file
        wb = openpyxl.load_workbook('Social Follower Count.xlsx', data_only=True)
        
        # Get data from all three sheets
        x_sheet = wb['X']
        discord_sheet = wb['Discord']
        telegram_sheet = wb['Telegram']  # Add Telegram sheet
        
        # Find token data in X sheet
        x_data = []
        for row in x_sheet.iter_rows(min_row=1):
            if row[0].value == token_name:
                for col in range(10, x_sheet.max_column, 2):
                    if row[col].value and row[col+1].value:
                        x_data.append({
                            'time': row[col].value,
                            'followers': row[col+1].value
                        })
                break
        
        # Find token data in Discord sheet
        discord_data = []
        for row in discord_sheet.iter_rows(min_row=1):
            if row[0].value == token_name:
                for col in range(10, discord_sheet.max_column, 2):
                    if row[col].value and row[col+1].value:
                        discord_data.append({
                            'time': row[col].value,
                            'followers': row[col+1].value
                        })
                break
        
        # Find token data in Telegram sheet
        telegram_data = []
        for row in telegram_sheet.iter_rows(min_row=1):
            if row[0].value == token_name:
                for col in range(10, telegram_sheet.max_column, 2):
                    if row[col].value and row[col+1].value:
                        telegram_data.append({
                            'time': row[col].value,
                            'followers': row[col+1].value
                        })
                break
                
        # Sort all data by time (newest first)
        x_data.sort(key=lambda x: x['time'], reverse=True)
        discord_data.sort(key=lambda x: x['time'], reverse=True)
        telegram_data.sort(key=lambda x: x['time'], reverse=True)
        
        return render_template('token_detail.html', 
                             token_name=token_name,
                             x_data=x_data,
                             discord_data=discord_data,
                             telegram_data=telegram_data)  # Add telegram data
                             
    except Exception as e:
        return f"Error: {str(e)}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)